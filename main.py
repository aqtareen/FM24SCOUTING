import sys
import os
import parser
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QGroupBox, QLineEdit, QComboBox, QPushButton, QLabel, QCheckBox, QFileDialog,
    QMessageBox, QScrollArea, QSpinBox, QDoubleSpinBox
)
from PyQt5.QtCore import Qt

# Position configuration with attributes and weights
POSITION_CONFIG = {
    "ST": {
        "attributes": {
            "Dribbling": 3, "Finishing": 5, "First Touch": 5, "Heading": 2.5, 
            "Long Shot": 2, "Passing": 3.5, "Technique": 3, "Aggression": 3.2,
            "Anticipation": 4.2, "Bravery": 3, "Composure": 4.2, "Concentration": 2.5,
            "Decisions": 3.2, "Determination": 4,"Flair": 1,"Leadership": 1.5, "Off the Ball": 5, "Teamwork": 2.5,
            "Vision": 3, "Work Rate": 3.5, "Acceleration": 4, "Agility": 3,
            "Balance": 3, "Jumping Reach": 1.5, "Natural Fitness": 2, "Pace": 3.5,
            "Stamina": 3.5, "Strength": 2.5
        },
        "filename": "ST_players.xlsx"
    },
    "ATTMID": {
        "attributes": {
            "Corners": 0, "Crossing": 1, "Dribbling": 3.5, "Finishing": 1,
            "First Touch": 4, "Heading": 1, "Long Shot": 3.5, "Passing": 5,
            "Technique": 4, "Aggression": 0, "Anticipation": 3.5, "Bravery": 1,
            "Composure": 3.7, "Concentration": 1, "Decisions": 4.5, "Determination": 1,
            "Flair": 4.5, "Leadership": 1.5, "Off the Ball": 5, "Positioning": 0,
            "Teamwork": 4, "Vision": 5, "Work Rate": 5, "Acceleration": 3.5,
            "Agility": 3.5, "Balance": 1, "Natural Fitness": 2, "Pace": 3,
            "Stamina": 5, "Strength": 1
        },
        "filename": "ATTMID_players.xlsx"
    },
    "CM": {
        "attributes": {
            "Crossing": 1.2, "Dribbling": 2, "Finishing": 1, "First Touch": 3,
            "Heading": 0, "Long Shot": 2, "Marking": 1, "Passing": 5,
            "Tackling": 2, "Technique": 2.2, "Aggression": 1, "Anticipation": 2.1,
            "Bravery": 1, "Composure": 2, "Concentration": 2.4, "Decisions": 4,
            "Determination": 0, "Flair": 1.5, "Leadership": 1.5, "Off the Ball": 2.5,
            "Positioning": 4, "Teamwork": 2.5, "Vision": 4, "Work Rate": 5,
            "Acceleration": 2.5, "Agility": 2.5, "Balance": 1, "Jumping Reach": 1,
            "Natural Fitness": 2, "Pace": 2.5, "Stamina": 5, "Strength": 1
        },
        "filename": "CM_players.xlsx"
    },
    "CDM": {
        "attributes": {
            "Crossing": 0.5, "Dribbling": 1, "Finishing": 0, "First Touch": 5,
            "Heading": 1, "Long Shot": 1.5, "Marking": 3, "Passing": 5,
            "Tackling": 4, "Technique": 1.5, "Aggression": 3, "Anticipation": 4.5,
            "Bravery": 1, "Composure": 5, "Concentration": 2.5, "Decisions": 2.7,
            "Determination": 1, "Flair": 0, "Leadership": 1.5, "Off the Ball": 1.6,
            "Positioning": 5, "Teamwork": 5, "Vision": 3, "Work Rate": 4,
            "Acceleration": 3.5, "Agility": 3.5, "Balance": 1.5, "Jumping Reach": 1.7,
            "Natural Fitness": 2, "Pace": 2.5, "Stamina": 2, "Strength": 3.5
        },
        "filename": "CDM_players.xlsx"
    },
    "CB": {
        "attributes": {
            "Dribbling": 0.1, "Finishing": 1.5, "First Touch": 1.2, "Heading": 1.4,
            "Long Shot": 0.17, "Marking": 2.3, "Passing": 2, "Tackling": 5,
            "Technique": 1.1, "Aggression": 1.3, "Anticipation": 1.4, "Bravery": 1.23,
            "Composure": 5, "Concentration": 4.8, "Decisions": 2.1, "Determination": 0.9,
            "Leadership": 1.5, "Off the Ball": 0, "Positioning": 2.5, "Teamwork": 1.75,
            "Vision": 2.2, "Work Rate": 2, "Acceleration": 3, "Agility": 2.8,
            "Balance": 0, "Jumping Reach": 5, "Natural Fitness": 2, "Pace": 3.5,
            "Stamina": 2, "Strength": 4.7
        },
        "filename": "CB_players.xlsx"
    },
    "WINGBACK": {
        "attributes": {
            "Crossing": 5, "Dribbling": 4, "Finishing": 1.2, "First Touch": 3.5,
            "Heading": 1.3, "Long Shot": 0.3, "Marking": 3, "Passing": 3.4,
            "Tackling": 2, "Technique": 4.9, "Aggression": 0.5, "Anticipation": 2.5,
            "Bravery": 1.5, "Composure": 1.5, "Concentration": 2.5, "Decisions": 3,
            "Determination": 4, "Flair": 1, "Leadership": 1.5, "Off the Ball": 3.8,
            "Positioning": 5, "Teamwork": 3.6, "Vision": 3.1, "Work Rate": 5,
            "Acceleration": 4, "Agility": 3.5, "Balance": 3.5, "Jumping Reach": 1,
            "Natural Fitness": 2, "Pace": 4, "Stamina": 5, "Strength": 1
        },
        "filename": "WINGBACK_players.xlsx"
    },
    "WINGER": {
        "attributes": {
            "Crossing": 2, "Dribbling": 5, "Finishing": 4, "First Touch": 3.5,
            "Heading": 0, "Long Shot": 1.25, "Passing": 3.5, "Technique": 4.8,
            "Anticipation": 3.5, "Composure": 3, "Concentration": 1, "Decisions": 3,
            "Determination": 1, "Flair": 4.5, "Leadership": 1.5, "Off the Ball": 3.8,
            "Positioning": 0, "Teamwork": 2.5, "Vision": 3, "Work Rate": 3.2,
            "Acceleration": 5, "Agility": 5, "Balance": 5, "Natural Fitness": 2,
            "Pace": 4.5, "Stamina": 3.5, "Strength": 1
        },
        "filename": "WINGER_players.xlsx"
    },
    "GK": {
        "attributes": {
            "Aerial Reach": 4.5, "Command of Area": 4, "Communication": 3, 
            "Eccentricity": -1, "First Touch": 2.5, "Handling": 3, "Kicking": 3.2,
            "One on Ones": 5, "Passing": 3.5, "Punching": 1, "Reflexes": 5, 
            "Rushing": 4, "Throwing": 3.2, "Aggression": -1, "Anticipation": 3.5, "Bravery": 1, 
            "Composure": 3, "Concentration": 3.5, "Decisions": 5, "Determination": 1,
            "Leadership": 1.5, "Positioning": 4, "Vision": 2, "Acceleration": 1,
            "Agility": 4.5, "Balance": 0, "Jumping Reach": 4.5, "Natural Fitness": 2,
            "Strength": 1
        },
        "filename": "GK_players.xlsx"
    }
}

class PlayerScoutApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Football Manager Player Scout")
        self.setGeometry(100, 100, 800, 800)
        
        # Central widget with scroll area
        central_widget = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(central_widget)
        scroll.setWidgetResizable(True)
        self.setCentralWidget(scroll)
        
        main_layout = QVBoxLayout(central_widget)
        
        # Player info group
        info_group = QGroupBox("Player Information")
        info_layout = QFormLayout()
        
        self.player_name = QLineEdit()
        self.position_combo = QComboBox()
        self.position_combo.addItems(POSITION_CONFIG.keys())
        self.position_combo.currentTextChanged.connect(self.update_attribute_fields)
        
        self.age_spin = QSpinBox()
        self.age_spin.setRange(15, 45)
        self.age_spin.setValue(25)
        
        # CHANGED: This checkbox now indicates if player is IN our club
        self.my_club_check = QCheckBox("Plays for my club")
        
        # Valuation should only be entered for players NOT in our club
        self.valuation_label = QLabel("Valuation (for players not in club):")
        self.valuation_edit = QDoubleSpinBox()
        self.valuation_edit.setRange(0, 200)
        self.valuation_edit.setPrefix("£")
        self.valuation_edit.setSuffix("m")
        self.valuation_edit.setEnabled(True)  # Enabled by default for non-club players
        
        # Connect checkbox to enable/disable valuation field
        self.my_club_check.toggled.connect(self.toggle_valuation_field)
        
        self.height_edit = QLineEdit()
        self.height_edit.setPlaceholderText("e.g., 6,1")
        
        self.weight_spin = QSpinBox()
        self.weight_spin.setRange(50, 100)
        self.weight_spin.setValue(75)
        self.weight_spin.setSuffix(" kg")
        
        self.verdict_edit = QLineEdit()
        self.verdict_edit.setPlaceholderText("Scouting notes...")
        
        info_layout.addRow("Player Name:", self.player_name)
        info_layout.addRow("Position:", self.position_combo)
        info_layout.addRow("Age:", self.age_spin)
        info_layout.addRow(self.my_club_check)
        info_layout.addRow(self.valuation_label, self.valuation_edit)
        info_layout.addRow("Height (ft,in):", self.height_edit)
        info_layout.addRow("Weight:", self.weight_spin)
        info_layout.addRow("Verdict:", self.verdict_edit)
        info_group.setLayout(info_layout)
        
        # Attributes container
        self.attributes_container = QGroupBox("Attributes")
        self.attributes_layout = QFormLayout()
        self.attributes_container.setLayout(self.attributes_layout)
        
        # Initialize attribute fields for first position
        self.update_attribute_fields(self.position_combo.currentText())
        
        # Buttons
        button_layout = QHBoxLayout()
        self.submit_btn = QPushButton("Submit Player")
        self.clear_btn = QPushButton("Clear Form")
        self.import_btn = QPushButton("Import Player")  
        
        button_layout.addWidget(self.submit_btn)
        button_layout.addWidget(self.clear_btn)
        button_layout.addWidget(self.import_btn)
        # Status label
        self.status_label = QLabel("Ready to scout players")
        self.status_label.setStyleSheet("color: gray; font-style: italic;")
        
        # Add widgets to main layout
        main_layout.addWidget(info_group)
        main_layout.addWidget(self.attributes_container)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.status_label)
        
        # Connect signals
        self.submit_btn.clicked.connect(self.submit_player)
        self.clear_btn.clicked.connect(self.clear_form)
        self.import_btn.clicked.connect(self.import_player)

    def import_player(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Open Player HTML", "", "HTML Files (*.html *.htm)"
        )
        
        if not filepath:
            return
            
        try:
            player_data = parser.parse_player_html(filepath)
            self.populate_form(player_data)
            self.status_label.setText("Player attributes imported successfully!")
        
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import player:\n{str(e)}")
            
    def populate_form(self, player_data):
        """Populate the form with imported player data"""
        # Safely handle height and weight
        if "height" in player_data:
            self.height_edit.setText(str(player_data["height"]))
        if "weight" in player_data:
            try:
                self.weight_spin.setValue(int(player_data["weight"]))
            except (TypeError, ValueError):
                pass  # Keep default if conversion fails

        # Set position if provided

        # Update attribute fields for the position
        position = self.position_combo.currentText()
        self.update_attribute_fields(position)

        # Set attribute values
        attributes = player_data.get("attributes", {})
        for i in range(self.attributes_layout.count()):
            item = self.attributes_layout.itemAt(i)
            if item.widget() and isinstance(item.widget(), QSpinBox):
                attr_name = item.widget().property("attribute")
                if attr_name and attr_name in attributes:
                    item.widget().setValue(attributes[attr_name])

                
    def toggle_valuation_field(self, checked):
        """Enable/disable valuation field based on club status"""
        # If player is in our club (checked), disable valuation field
        # If player is NOT in our club (unchecked), enable valuation field
        self.valuation_edit.setEnabled(not checked)
        
        # Clear valuation if player is in our club
        if checked:
            self.valuation_edit.setValue(0)
    
    def update_attribute_fields(self, position):
        """Update attribute fields based on selected position"""
        # Clear existing fields
        for i in reversed(range(self.attributes_layout.count())):
            widget = self.attributes_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        # Add new fields for position
        config = POSITION_CONFIG.get(position, {})
        attributes = config.get("attributes", {})
        
        for attr, weight in attributes.items():
            if weight == 0:  # Skip attributes with 0 weight
                continue
                
            spin = QSpinBox()
            spin.setRange(1, 20)
            spin.setValue(10)
            spin.setProperty("attribute", attr)
            spin.setProperty("weight", weight)
            self.attributes_layout.addRow(f"{attr} (×{weight}):", spin)
    
    def clear_form(self):
        """Reset all form fields"""
        self.player_name.clear()
        self.position_combo.setCurrentIndex(0)
        self.age_spin.setValue(25)
        self.my_club_check.setChecked(False)
        self.valuation_edit.setValue(0)
        self.valuation_edit.setEnabled(True)  # Enable by default
        self.height_edit.clear()
        self.weight_spin.setValue(75)
        self.verdict_edit.clear()
        
        # Reset attribute fields
        for i in range(self.attributes_layout.count()):
            item = self.attributes_layout.itemAt(i)
            if item.widget() and isinstance(item.widget(), QSpinBox):
                item.widget().setValue(10)
        
        self.status_label.setText("Form cleared")
    
    def get_attribute_values(self):
        """Get all attribute values and weights from form"""
        attributes = {}
        for i in range(self.attributes_layout.count()):
            item = self.attributes_layout.itemAt(i)
            if item.widget() and isinstance(item.widget(), QSpinBox):
                attr = item.widget().property("attribute")
                weight = item.widget().property("weight")
                value = item.widget().value()
                attributes[attr] = {"value": value, "weight": weight}
        return attributes
    
    def calculate_score(self, position):
        """Calculate weighted score for the player"""
        total_score = 0
        attributes = self.get_attribute_values()
        
        for attr_data in attributes.values():
            value = attr_data["value"]
            weight = attr_data["weight"]
            total_score += value * weight
        
        return round(total_score, 2)
    
    def create_position_workbook(self, position):
        """Create a new Excel workbook for a position"""
        config = POSITION_CONFIG[position]
        wb = Workbook()
        ws = wb.active
        ws.title = "Players"
        
        # Create headers
        headers = ["Name", "Position", "Age", "Valuation", "Height", "Weight", "Total Score"]
        headers.extend(config["attributes"].keys())
        headers.append("Verdict")
        
        ws.append(headers)
        return wb
    
    def save_player(self, position):
        """Save player data to position-specific Excel file"""
        config = POSITION_CONFIG[position]
        filename = config["filename"]
        
        try:
            # Check if file exists
            if os.path.exists(filename):
                wb = load_workbook(filename)
                ws = wb.active
            else:
                wb = self.create_position_workbook(position)
                ws = wb.active
            
            # Determine valuation based on club status
            if self.my_club_check.isChecked():
                valuation = "N/A"  # Player is in our club, no valuation needed
            else:
                valuation = self.valuation_edit.value()
            
            # Prepare player data
            player_data = [
                self.player_name.text(),
                position,
                self.age_spin.value(),
                valuation,
                self.height_edit.text(),
                self.weight_spin.value(),
                self.calculate_score(position)
            ]
            
            # Add attribute values in the correct order
            attributes = self.get_attribute_values()
            for attr in config["attributes"]:
                if attr in attributes:
                    player_data.append(attributes[attr]["value"])
                else:
                    player_data.append("")  # For attributes with 0 weight
            
            player_data.append(self.verdict_edit.text())
            
            # Add new row
            ws.append(player_data)
            
            # Save workbook
            wb.save(filename)
            return True, filename
        except Exception as e:
            return False, str(e)
    
    def submit_player(self):
        """Handle player submission"""
        # Validate inputs
        if not self.player_name.text().strip():
            QMessageBox.warning(self, "Missing Information", "Player name is required")
            return
        
        position = self.position_combo.currentText()
        if position not in POSITION_CONFIG:
            QMessageBox.warning(self, "Error", "Invalid position selected")
            return
        
        # Validate valuation for players not in club
        if not self.my_club_check.isChecked() and self.valuation_edit.value() <= 0:
            QMessageBox.warning(self, "Missing Valuation", 
                               "Valuation is required for players not in your club")
            return
        
        # Save player
        success, result = self.save_player(position)
        
        if success:
            self.status_label.setText(f"Player saved to {result}")
            self.clear_form()
            QMessageBox.information(self, "Success", "Player data saved successfully!")
        else:
            QMessageBox.critical(self, "Error", f"Failed to save player:\n{result}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PlayerScoutApp()
    window.show()
    sys.exit(app.exec_())